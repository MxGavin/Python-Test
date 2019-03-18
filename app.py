print("Hello World!")
import math
print( pow(2,3))
#
# from Book import Book
# book1 =  Book("Harry Potter", "JK Rowling");
# print(book1.title)
# book1.read_book()
#
# print('*' * 10)
#
# course = 'Python for "Beginners"'
# course1 = '''
# Hi John,
#
# Here is our first email to you.
#
# Thank you,
# The support team
#
# '''
# print(course)
# print(course1)
#
# first = 'John'
# last = 'Smith'
# message = f'{first} [{last}] is a coder'
# print(message)
#
# print('Python' in course)

#
# command = ""
# is_started = False
# while True:
#     command = input("> ").lower()
#     if command == "start":
#         if is_started:
#             print("Car is already started!")
#         else:
#             is_started = True
#             print("Car started...")
#     elif command == "stop":
#         if not is_started:
#             print("Car is already stopped.")
#         else:
#             is_started = False
#             print("Car stopped.")
#     elif command == "help":
#         print("""
# start - to start the car
# stop - to stop the car
# quit - to quit
#         """)
#     elif command == "quit":
#         break
#     else:
#         print("Sorry, I don't understand that!")
#

# message = input(">")
# words = message.split(' ')
# emojis = {
#     ":)": "Happy",
#     ":(": "Sad"
# }
#
# output = ""
# for word in words:
#     output += emojis.get(word, word) + " "
#
# print(output)

from ecommerce.shipping import calc_shipping
calc_shipping()

# import random
# members = ["Mary", "Gavin", "Ethan"]
# for i in range(3):
#     print(random.randint(10,20))
#
# print(random.choice(members))
#
#
# class Dice:
#     def roll(self):
#         first = random.randint(1,6)
#         second = random.randint(1,6)
#         return first, second
#
#
# dice = Dice()
# print(dice.roll())

# from pathlib import Path
# # path = Path("ecommerce")
# # print(path.exists())
#
# path_junior = Path()
# for file in path_junior.glob('*.py'):
#     print(file)
#
#


import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# wb = xl.load_workbook('transactions.xlsx')
# sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)
# print(cell.value)

#
# def process_workbook(filename):
#     wb = xl.load_workbook(filename)
#     sheet = wb['Sheet1']
#
#     for row in range(2, sheet.max_row + 1):
#         cell = sheet.cell(row, 3)
#         corrected_price = cell.value * 0.9
#         corrected_price_cell = sheet.cell(row, 4)
#         corrected_price_cell.value = corrected_price
#
#     values = Reference(sheet,
#                        min_row=2,
#                        max_row=sheet.max_row,
#                        min_col=4,
#                        max_col=4)
#
#     chart = BarChart()
#     chart.add_data(values)
#     sheet.add_chart(chart, 'E2')
#
#     wb.save(filename)
#
#
# process_workbook('transactions.xlsx')
#


girls = ['alice', 'bernice', 'clarice']
boys = ['chris', 'arnold', 'bob']
# bg_list = [b+'+'+g for b in boys for g in girls if b[0]==g[0]]
print([b+'+'+g for b in boys for g in girls if b[0]==g[0]])

letterGirls = {}
for girl in girls:
    letterGirls.setdefault(girl[0], []).append(girl)
print([b+'+'+g for b in boys for g in letterGirls[b[0]]])
# print(len(scope))

for b, g in zip(boys, girls):
    print(f'{b} and {g}')

nums = [1,2,3]

for index, name in enumerate(boys):
    if "c" in name:
        boys[index] = 'Catch you'
print(boys)


# mylist = [n*n for n in nums]
# print(mylist)

# mylist = map(lambda n: n*n, nums)
# print(mylist)
